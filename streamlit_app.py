import streamlit as st
import pandas as pd
from textblob import TextBlob
import plotly.express as px
import plotly.graph_objects as go
from PyPDF2 import PdfReader
from fpdf import FPDF
import openpyxl
import io

st.set_page_config(page_title="Product Review Sentiment Analysis", page_icon="💬", layout="wide")

st.title("💬 Product Review Sentiment Analysis")
st.markdown("Analyze the sentiment of product reviews to understand customer feedback")

tab1, tab2, tab3, tab4 = st.tabs(["📁 Bulk Analysis", "📊 Dashboard", "📝 Review Analysis", "❓ Help Center"])

with tab1:
    st.header("Bulk Review Analysis")
    
    input_method = st.radio("Choose input method:", ["Text Input", "CSV Upload", "PDF Upload"])
    
    if input_method == "Text Input":
        bulk_text = st.text_area("Enter multiple reviews (one per line):", height=200)
        
        if st.button("Analyze All Reviews", type="primary"):
            if bulk_text.strip():
                reviews = [r.strip() for r in bulk_text.split('\n') if r.strip()]
                results = []
                
                for review in reviews:
                    blob = TextBlob(review)
                    polarity = blob.sentiment.polarity
                    subjectivity = blob.sentiment.subjectivity
                    
                    if polarity > 0.1:
                        sentiment = "Positive"
                    elif polarity < -0.1:
                        sentiment = "Negative"
                    else:
                        sentiment = "Neutral"
                    
                    results.append({
                        'review': review,
                        'sentiment': sentiment,
                        'polarity': polarity,
                        'subjectivity': subjectivity
                    })
                
                df = pd.DataFrame(results)
                st.session_state.reviews_df = df
                
                st.success(f"Analyzed {len(reviews)} reviews!")
                st.dataframe(df, use_container_width=True)
            else:
                st.warning("Please enter some reviews.")
    
    elif input_method == "PDF Upload":
        st.subheader("Upload Product Review PDF")
        
        uploaded_pdf = st.file_uploader("Upload a PDF file containing reviews:", type=['pdf'])
        
        if uploaded_pdf is not None:
            try:
                pdf_reader = PdfReader(uploaded_pdf)
                extracted_text = ""
                
                for page in pdf_reader.pages:
                    extracted_text += page.extract_text() + "\n"
                
                st.success(f"Extracted {len(pdf_reader.pages)} pages from PDF")
                
                st.text_area("Extracted Text:", extracted_text, height=200)
                
                if st.button("Analyze PDF Content", type="primary"):
                    reviews = [r.strip() for r in extracted_text.split('\n') if r.strip() and len(r.strip()) > 10]
                    results = []
                    
                    for review in reviews:
                        blob = TextBlob(review)
                        polarity = blob.sentiment.polarity
                        subjectivity = blob.sentiment.subjectivity
                        
                        if polarity > 0.1:
                            sentiment = "Positive"
                        elif polarity < -0.1:
                            sentiment = "Negative"
                        else:
                            sentiment = "Neutral"
                        
                        results.append({
                            'review': review,
                            'sentiment': sentiment,
                            'polarity': polarity,
                            'subjectivity': subjectivity
                        })
                    
                    df = pd.DataFrame(results)
                    st.session_state.reviews_df = df
                    
                    st.success(f"Analyzed {len(reviews)} reviews from PDF!")
                    st.dataframe(df, use_container_width=True)
                    
                    csv = df.to_csv(index=False)
                    st.download_button("Download Results", csv, "sentiment_analysis_results.csv", "text/csv")
                    
            except Exception as e:
                st.error(f"Error reading PDF: {str(e)}")
    
    else:
        uploaded_file = st.file_uploader("Upload CSV file with reviews:", type=['csv'])
        
        if uploaded_file is not None:
            df_upload = pd.read_csv(uploaded_file)
            st.write("Preview of uploaded data:")
            st.dataframe(df_upload.head())
            
            review_column = st.selectbox("Select the column containing reviews:", df_upload.columns)
            
            if st.button("Analyze CSV Reviews", type="primary"):
                results = []
                
                for review in df_upload[review_column]:
                    if pd.notna(review):
                        blob = TextBlob(str(review))
                        polarity = blob.sentiment.polarity
                        subjectivity = blob.sentiment.subjectivity
                        
                        if polarity > 0.1:
                            sentiment = "Positive"
                        elif polarity < -0.1:
                            sentiment = "Negative"
                        else:
                            sentiment = "Neutral"
                        
                        results.append({
                            'review': review,
                            'sentiment': sentiment,
                            'polarity': polarity,
                            'subjectivity': subjectivity
                        })
                
                df = pd.DataFrame(results)
                st.session_state.reviews_df = df
                
                st.success(f"Analyzed {len(results)} reviews!")
                st.dataframe(df, use_container_width=True)
                
                csv = df.to_csv(index=False)
                st.download_button("Download Results", csv, "sentiment_analysis_results.csv", "text/csv")

with tab2:
    st.header("Sentiment Dashboard")
    
    if 'reviews_df' in st.session_state and not st.session_state.reviews_df.empty:
        df = st.session_state.reviews_df

        st.subheader("Filters and Sorting")
        col1, col2, col3 = st.columns(3)
        with col1:
            sentiments = st.multiselect("Filter by Sentiment", options=['Positive', 'Neutral', 'Negative'], default=['Positive', 'Neutral', 'Negative'])
        with col2:
            polarity_range = st.slider("Polarity Range", min_value=-1.0, max_value=1.0, value=(-1.0, 1.0))
        with col3:
            sort_by = st.selectbox("Sort by", options=['polarity', 'subjectivity'], index=0)

        filter_df = df[df['sentiment'].isin(sentiments) & (df['polarity'] >= polarity_range[0]) & (df['polarity'] <= polarity_range[1])]
        if sort_by == 'polarity':
            filter_df = filter_df.sort_values('polarity', ascending=False)
        else:
            filter_df = filter_df.sort_values('subjectivity', ascending=False)

        col1, col2, col3 = st.columns(3)

        positive_count = (filter_df['polarity'] > 0.1).sum()
        neutral_count = ((filter_df['polarity'] >= -0.1) & (filter_df['polarity'] <= 0.1)).sum()
        negative_count = (filter_df['polarity'] < -0.1).sum()

        with col1:
            st.metric("Positive Reviews", positive_count, delta=f"{positive_count/len(filter_df)*100:.1f}%" if len(filter_df) > 0 else "0%")
        with col2:
            st.metric("Neutral Reviews", neutral_count, delta=f"{neutral_count/len(filter_df)*100:.1f}%" if len(filter_df) > 0 else "0%")
        with col3:
            st.metric("Negative Reviews", negative_count, delta=f"{negative_count/len(filter_df)*100:.1f}%" if len(filter_df) > 0 else "0%")

        fig_pie = px.pie(
            filter_df,
            names='sentiment',
            title='Sentiment Distribution',
            color='sentiment',
            color_discrete_map={
                'Positive': '#28a745',
                'Neutral': '#6c757d',
                'Negative': '#dc3545'
            }
        )
        st.plotly_chart(fig_pie, use_container_width=True)

        fig_hist = px.histogram(
            filter_df,
            x='polarity',
            nbins=20,
            title='Polarity Score Distribution',
            labels={'polarity': 'Polarity Score', 'count': 'Count'}
        )
        st.plotly_chart(fig_hist, use_container_width=True)

        # Additional visualizations
        fig_scatter = px.scatter(
            filter_df,
            x='polarity',
            y='subjectivity',
            color='sentiment',
            title='Polarity vs Subjectivity',
            labels={'polarity': 'Polarity', 'subjectivity': 'Subjectivity'}
        )
        st.plotly_chart(fig_scatter, use_container_width=True)

        fig_bar = px.bar(
            filter_df.groupby('sentiment').size().reset_index(name='count'),
            x='sentiment',
            y='count',
            title='Sentiment Counts',
            color='sentiment',
            color_discrete_map={
                'Positive': '#28a745',
                'Neutral': '#6c757d',
                'Negative': '#dc3545'
            }
        )
        st.plotly_chart(fig_bar, use_container_width=True)

        st.subheader("Review Details")
        st.dataframe(filter_df[['review', 'sentiment', 'polarity', 'subjectivity']], use_container_width=True)

        st.subheader("Export Reports")
        col1, col2, col3 = st.columns(3)
        with col1:
            csv = df.to_csv(index=False)
            st.download_button("Download CSV", csv, "sentiment_report.csv", "text/csv")
        with col2:
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Sentiment Analysis', index=False)
            buffer.seek(0)
            st.download_button("Download Excel", buffer, "sentiment_report.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with col3:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            pdf.cell(200, 10, txt="Sentiment Analysis Report", ln=True, align='C')
            pdf.ln(10)
            pdf.cell(200, 10, txt=f"Total Reviews: {len(df)}", ln=True)
            pdf.cell(200, 10, txt=f"Positive: {positive_count}, Neutral: {neutral_count}, Negative: {negative_count}", ln=True)
            pdf.ln(10)
            for index, row in df.head(20).iterrows():  # Limit to first 20 for PDF
                pdf.cell(200, 10, txt=f"Sentiment: {row['sentiment']} - Polarity: {row['polarity']:.2f}", ln=True)
                pdf.multi_cell(0, 10, txt=f"Review: {row['review'][:100]}...")
                pdf.ln(5)
            pdf_output = pdf.output(dest='S').encode('latin1')
            st.download_button("Download PDF", pdf_output, "sentiment_report.pdf", "application/pdf")
    else:
        st.info("No reviews analyzed yet. Go to 'Bulk Analysis' to upload reviews.")

with tab3:
    st.header("Review Analysis")
    
    review_text = st.text_area("Enter your product review:", height=150, placeholder="Type or paste a product review here...")
    
    if st.button("Analyze Sentiment", type="primary"):
        if review_text.strip():
            blob = TextBlob(review_text)
            polarity = blob.sentiment.polarity
            subjectivity = blob.sentiment.subjectivity
            
            if polarity > 0.1:
                sentiment = "Positive 😊"
                color = "green"
            elif polarity < -0.1:
                sentiment = "Negative 😞"
                color = "red"
            else:
                sentiment = "Neutral 😐"
                color = "gray"
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Sentiment", sentiment)
            with col2:
                st.metric("Polarity Score", f"{polarity:.3f}", delta=f"{polarity:.3f}", delta_color="normal")
            with col3:
                st.metric("Subjectivity", f"{subjectivity:.3f}")
            
            st.progress((polarity + 1) / 2)
            
            st.subheader("Analysis Details")
            st.write(f"**Polarity** ranges from -1 (negative) to +1 (positive)")
            st.write(f"**Subjectivity** ranges from 0 (objective) to 1 (subjective)")
            
            words = blob.words
            if words:
                st.write("**Key Words:**", ", ".join([str(w) for w in words[:10]]))

            st.subheader("AI-Powered Insights and Recommendations")
            if polarity > 0.5:
                st.success("🎉 This review is highly positive! Your product is doing great. Consider highlighting customer testimonials.")
            elif polarity > 0.1:
                st.info("👍 Positive feedback. Keep up the good work!")
            elif polarity > -0.1:
                st.warning("😐 Neutral sentiment. There might be room for improvement in engagement.")
            elif polarity > -0.5:
                st.error("👎 Negative feedback detected. Address concerns promptly to improve customer satisfaction.")
            else:
                st.error("😡 Strongly negative review. Immediate action recommended: review product features or customer service.")

            # Simple keyword-based recommendations
            negative_keywords = ['bad', 'worst', 'hate', 'terrible', 'disappointed']
            if any(kw in review_text.lower() for kw in negative_keywords):
                st.write("**Recommendation:** Focus on quality control and customer support.")
        else:
            st.warning("Please enter a review to analyze.")

with tab4:
    st.header("Help Center")

    with st.expander("Tutorials"):
        st.write("### How to Use Bulk Analysis")
        st.write("1. Choose input method: Text, CSV, or PDF.")
        st.write("2. Upload or enter reviews.")
        st.write("3. Click 'Analyze' to get sentiment results.")
        st.write("### Dashboard Filters")
        st.write("Use filters to narrow down reviews by sentiment and polarity range.")
        st.write("### Exporting Reports")
        st.write("In Dashboard, download reports in CSV, Excel, or PDF formats.")

    with st.expander("FAQs"):
        st.write("**Q: What is sentiment analysis?**")
        st.write("A: Sentiment analysis determines if text is positive, negative, or neutral.")
        st.write("**Q: How accurate is the analysis?**")
        st.write("A: Based on TextBlob, it's a basic model. For better accuracy, consider advanced AI models.")
        st.write("**Q: Is my data secure?**")
        st.write("A: Reviews are processed locally and not stored permanently.")

    st.subheader("Live Chat Support")
    st.write("For live support, contact us at support@sentimentapp.com or use the form below.")

    with st.form("contact_form"):
        name = st.text_input("Name")
        email = st.text_input("Email")
        message = st.text_area("Message")
        submitted = st.form_submit_button("Send")
        if submitted:
            st.success("Message sent! We'll get back to you soon.")

st.sidebar.header("About")
st.sidebar.info("""
This app uses **TextBlob** for sentiment analysis.

**Metrics:**
- **Polarity**: -1 to +1 (negative to positive)
- **Subjectivity**: 0 to 1 (objective to subjective)
""")

# Dark mode toggle
theme = st.sidebar.selectbox("Choose Theme", ["Light", "Dark"])
if theme == "Dark":
    st.markdown("""
    <style>
    body, .stApp { background-color: #121212; color: white; }
    .css-1d391kg { background-color: #121212; }
    </style>
    """, unsafe_allow_html=True)